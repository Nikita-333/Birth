import os
import time
from datetime import datetime, timedelta

import openpyxl
import plyer
import schedule
from plyer import notification
from tabulate import tabulate
from apscheduler.schedulers.background import BackgroundScheduler

MAX_MESSAGE_LENGTH = 256


def read_excel_table():
    workbook = openpyxl.load_workbook("Birth.xlsx")
    worksheet = workbook.active

    data = []
    headers = [cell.value for cell in worksheet[1]]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    table = tabulate(data, headers, tablefmt="grid")
    return table


def add_employee_to_excel(file_name, full_name, date_of_birth):
    try:
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.append(["Имя сотрудника или кандидата", "Дата рождения"])

    try:
        from datetime import datetime
        date_of_birth = datetime.strptime(date_of_birth, "%m/%d/%Y").strftime("%m/%d/%Y")
    except ValueError:
        return f"Ошибка: Некорректный формат даты. Введите дату в формате MM/DD/YY."

    worksheet.append([full_name, date_of_birth])

    workbook.save(file_name)
    return f"Сотрудник успешно добавлен."


def delete_employee_by_name(file_name, name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    found = False

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == name:
            worksheet.delete_rows(row_index, 1)
            found = True
            break

    if found:
        workbook.save(file_name)
        return f"Строка с именем '{name}' успешно удалена."
    else:
        return f"Строка с именем '{name}' не найдена."


def find_nearest_birthday(file_name):
    current_date = datetime.today().date()

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    birthdays = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        birthday = row[1]
        if birthday is not None:
            birthday = datetime.strptime(birthday, "%m/%d/%Y").date().replace(year=current_date.year)
            birthdays.append(birthday)

    birthdays.sort()

    next_birthday = None
    result = ""

    for birthday in birthdays:
        if birthday >= current_date:
            next_birthday = birthday
            break

    if next_birthday:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] and datetime.strptime(row[1], "%m/%d/%Y").date().replace(year=current_date.year) == next_birthday:
                name = row[0]
                days_difference = (next_birthday - current_date).days
                result += f"Имя сотрудника или кандидата: {name}\n"
                result += f"Ближайший день рождения: {next_birthday.strftime('%m/%d/%Y')}\n"
                result += f"Дата рождения: {datetime.strptime(row[1], '%m/%d/%Y').date().strftime('%m/%d/%Y')}\n"
                result += f"Разница в днях: {days_difference}\n\n"

    if result:
        return result.rstrip()
    else:
        return None


def find_next_birthday(file_name):
    current_date = datetime.today().date()

    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    birthdays = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        birthday = row[1]
        if birthday is not None:
            birthday = datetime.strptime(birthday, "%m/%d/%Y").date().replace(year=current_date.year)
            birthdays.append(birthday)

    birthdays.sort()

    next_birthday = None
    next_birthday_found = False

    for birthday in birthdays:
        if birthday >= current_date:
            if next_birthday_found:
                next_birthday = birthday
                break
            else:
                next_birthday_found = True

    if next_birthday:
        result = ""

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] and datetime.strptime(row[1], "%m/%d/%Y").date().replace(year=current_date.year) == next_birthday:
                name = row[0]
                days_difference = (next_birthday - current_date).days
                result += f"Имя сотрудника или кандидата: {name}\n"
                result += f"Ближайший день рождения: {next_birthday.strftime('%m/%d/%Y')}\n"
                result += f"Дата рождения: {datetime.strptime(row[1], '%m/%d/%Y').date().strftime('%m/%d/%Y')}\n"
                result += f"Разница в днях: {days_difference}\n\n"

        if result:
            return result.rstrip()

    return None


# datetime.strptime(row[1], "%m/%d/%Y").date().strftime("%m/%d/%Y")
def find_birthday_by_lastname(file_name, name_or_lastname):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active

    employees_found = False
    result = ""

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        full_name = row[0]
        birthday = row[1]

        if full_name is not None:
            if name_or_lastname.lower() in full_name.lower():
                if not employees_found:
                    result += f"Сотрудники с фамилией или именем {name_or_lastname}:\n"
                    employees_found = True

                if birthday is not None:
                    result += f"Имя сотрудника или кандидата: {full_name}\n"
                    result += f"Дата рождения: {birthday}\n\n"
                else:
                    result += f"Дата рождения для сотрудника '{full_name}' неизвестна.\n"

    if not employees_found:
        result += f"Сотрудники с фамилией или именем '{name_or_lastname}' не найдены.\n"

    return result

def check_first_run():
    if os.path.exists("first_run.txt"):
        return False
    else:
        with open("first_run.txt", "w"):
            pass
        return True

def notify_birthday():
    today = datetime.today().date()
    tomorrow = today + timedelta(days=1)
    file_name = "Birth.xlsx"
    nearest_birth = find_nearest_birthday(file_name)
    next_birth = find_next_birthday(file_name)

    if today.strftime("%m/%d/%Y") in nearest_birth:
        employees = nearest_birth.split("\n\n")
        for i, employee in enumerate(employees):
            send_long_notification(f"День рождения сегодня ({i + 1}/{len(employees)})", employee)

    if tomorrow.strftime("%m/%d/%Y") in next_birth:
        employeesj = next_birth.split("\n\n")
        for j, employeej in enumerate(employeesj):
            send_long_notification(f"День рождения завтра ({j + 1}/{len(employeesj)})", employeej)


def send_long_notification(title, message):
    if len(message) <= MAX_MESSAGE_LENGTH:
        notification.notify(
            title=title,
            message=message,
            app_icon=None,
            timeout=10
        )
    else:
        chunks = [message[i:i + MAX_MESSAGE_LENGTH] for i in range(0, len(message), MAX_MESSAGE_LENGTH)]
        num_chunks = len(chunks)

        for i, chunk in enumerate(chunks):
            notification.notify(
                title=f"{title} ({i + 1}/{num_chunks})",
                message=chunk,
                app_icon=None,
                timeout=10
            )

def run_scheduler():
    if check_first_run():
        file_name = "Birth.xlsx"
        scheduler = BackgroundScheduler()
        scheduler.add_job(notify_birthday, 'cron', hour=7, minute=0)
        scheduler.start()

        try:
            while True:
                pass
        except (KeyboardInterrupt, SystemExit):
            scheduler.shutdown()

#notify_birthday()

#run_scheduler()
